#!/usr/bin/env python3
"""Build a sterol platemap with biological replicates 1-3.

Workflow:
1. Use an existing platemap with biological replicate 1 as template.
2. Pull replicate 2 and 3 values from two Results workbooks.
3. Merge metric values by `Media + Strain`.
4. Keep `%` without total, and add totals only for raw / NORM / g per mg.
"""

from __future__ import annotations

import argparse
import numbers
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_BASE = Path(r"C:/Users/user/OneDrive/Escritorio/platemap_Ergosterol_Tot.xlsx")
DEFAULT_REP2 = Path(r"C:/Users/user/Downloads/Results260320.xlsx")
DEFAULT_REP3 = Path(r"C:/Users/user/Downloads/Results260330.xlsx")
DEFAULT_OUTPUT = Path(
    r"C:/Users/user/OneDrive/Escritorio/platemap_Ergosterol_Tot_rep123.xlsx"
)

META_COLS = [
    "Well",
    "Strain",
    "Media",
    "Orden",
    "Replicate",
    "BiologicalReplicate",
    "TechnicalReplicate",
]

BASE_COMPOUNDS = [
    "Ergosta-5_8_22_24(28)-tetraenol",
    "Zymosterol",
    "Lichesterol",
    "Ergosterol",
    "Ergosta-8_22-dienol",
    "Episterol",
    "Ergosta-5_7_24(28)-trienol",
    "Ergosta-7_22-dienol",
    "Ergosta-7-enol ",
    "FF-MAS",
    "Lanosterol",
    "T-MAS",
    "Two unkown diols",
]

TOTAL_NAME = "Total sterols"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create a new ergosterol platemap with biological replicates 1-3 "
            "using one base platemap (rep1) and two Results files (rep2, rep3)."
        )
    )
    parser.add_argument(
        "--base",
        default=str(DEFAULT_BASE),
        help="Template platemap .xlsx containing replicate 1 in sheet Datos.",
    )
    parser.add_argument(
        "--rep2",
        default=str(DEFAULT_REP2),
        help="Results workbook used as biological replicate 2.",
    )
    parser.add_argument(
        "--rep3",
        default=str(DEFAULT_REP3),
        help="Results workbook used as biological replicate 3.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Output platemap .xlsx path.",
    )
    return parser.parse_args()


def ensure_exists(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{label} not found: {path}")


def backup_if_exists(path: Path) -> Path | None:
    if not path.exists():
        return None
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.with_name(f"{path.stem}.backup_{timestamp}{path.suffix}")
    shutil.copy2(path, backup)
    return backup


def normalize_base_strain(value: Any) -> str:
    if value is None:
        return "NA"
    if isinstance(value, float) and pd.isna(value):
        return "NA"
    text = str(value).strip()
    if not text:
        return "NA"
    return "NA" if text.upper() == "NA" else text


def normalize_sample(sample: Any) -> str:
    text = "" if sample is None else str(sample).strip()
    if not text:
        raise ValueError("Sample value is empty.")
    if text.upper() == "NA":
        return "NA"
    if text.startswith("BY4742"):
        suffix = text[len("BY4742") :].strip()
        if suffix == "":
            return "BY4742"
        return suffix
    return text


def normalize_media(raw_treatment: Any) -> str:
    if raw_treatment is None:
        raise ValueError("Missing treatment label in source row.")
    text = str(raw_treatment).strip().lower()
    if "co-treatment" in text or ("rapamycin" in text and "u18666a" in text):
        return "Rapa-U18"
    if "mock" in text:
        return "Mock"
    if "rapamycin" in text:
        return "Rapa"
    if "u18666a" in text:
        return "U18"
    raise ValueError(f"Unsupported treatment label: {raw_treatment!r}")


def find_header_row(ws: Worksheet) -> int:
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        val = ws.cell(row_idx, 1).value
        if isinstance(val, str) and val.strip().lower() == "sample id":
            return row_idx
    raise ValueError(f"Could not find 'Sample ID' header in sheet {ws.title!r}.")


def parse_results_sheet(ws: Worksheet, group: str) -> dict[tuple[str, str], dict[str, Any]]:
    header_row = find_header_row(ws)
    current_treatment: Any = None
    rows: dict[tuple[str, str], dict[str, Any]] = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        sample_id = ws.cell(row_idx, 1).value
        sample_name = ws.cell(row_idx, 2).value
        treatment = ws.cell(row_idx, 3).value

        if treatment is not None and str(treatment).strip() != "":
            current_treatment = treatment

        if not isinstance(sample_id, numbers.Number):
            continue
        if sample_name is None or str(sample_name).strip() == "":
            continue

        media = normalize_media(current_treatment)
        strain = normalize_sample(sample_name)
        key = (media, strain)

        values: dict[str, Any] = {}
        for idx, compound in enumerate(BASE_COMPOUNDS, start=6):
            values[compound] = ws.cell(row_idx, idx).value
        if group != "pct":
            values[TOTAL_NAME] = ws.cell(row_idx, 19).value

        if key in rows:
            raise ValueError(
                f"Duplicate key {key} found in sheet {ws.title!r} at row {row_idx}."
            )
        rows[key] = values

    return rows


def locate_metric_sheets(workbook_path: Path) -> dict[str, str]:
    wb = load_workbook(workbook_path, data_only=True)
    mapping: dict[str, str] = {}
    for sheet_name in wb.sheetnames:
        low = sheet_name.strip().lower()
        if low == "%":
            mapping["pct"] = sheet_name
        elif "raw" in low:
            mapping["raw"] = sheet_name
        elif "normal" in low:
            mapping["norm"] = sheet_name
        elif "per mg" in low:
            mapping["ug"] = sheet_name

    required = {"raw", "norm", "ug", "pct"}
    missing = sorted(required - set(mapping))
    if missing:
        raise ValueError(
            f"Workbook {workbook_path} is missing required metric sheets: {missing}"
        )
    return mapping


def parse_results_workbook(path: Path) -> dict[tuple[str, str], dict[str, dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    sheet_map = locate_metric_sheets(path)

    merged: dict[tuple[str, str], dict[str, dict[str, Any]]] = {}
    for group, sheet_name in sheet_map.items():
        per_sheet = parse_results_sheet(wb[sheet_name], group=group)
        for key, values in per_sheet.items():
            merged.setdefault(key, {})[group] = values
    return merged


def build_metric_column_map(columns: list[str]) -> dict[str, tuple[str, str]]:
    metric_map: dict[str, tuple[str, str]] = {}
    for col in columns:
        if col in META_COLS:
            continue
        if col in BASE_COMPOUNDS or col == TOTAL_NAME:
            metric_map[col] = ("raw", col)
            continue
        if col.startswith("NORM "):
            metric_map[col] = ("norm", col[len("NORM ") :])
            continue
        if col.startswith("% "):
            metric_map[col] = ("pct", col[len("% ") :])
            continue
        if "g/mg " in col:
            metric_map[col] = ("ug", col.split("g/mg ", 1)[1])
            continue
        raise ValueError(f"Unrecognized metric column in base platemap: {col!r}")
    return metric_map


def build_output(
    base_path: Path,
    rep2_path: Path,
    rep3_path: Path,
    output_path: Path,
) -> None:
    base_datos = pd.read_excel(base_path, sheet_name="Datos", dtype=object)
    base_plot = pd.read_excel(base_path, sheet_name="PlotSettings", dtype=object)
    ordered_cols = list(base_datos.columns)
    metric_map = build_metric_column_map(ordered_cols)

    rep2_data = parse_results_workbook(rep2_path)
    rep3_data = parse_results_workbook(rep3_path)
    source_by_bio = {2: rep2_data, 3: rep3_data}

    base_keys = {
        (
            str(row["Media"]).strip(),
            normalize_base_strain(row["Strain"]),
        )
        for _, row in base_datos.iterrows()
    }

    output_rows: list[dict[str, Any]] = []
    well_idx = 1
    replicated_counts = {1: 0, 2: 0, 3: 0}

    for _, base_row in base_datos.iterrows():
        key = (
            str(base_row["Media"]).strip(),
            normalize_base_strain(base_row["Strain"]),
        )

        for bio_rep in (1, 2, 3):
            if bio_rep == 1:
                row_data = {col: base_row[col] for col in ordered_cols}
            else:
                source = source_by_bio[bio_rep]
                metric_values = source.get(key)
                if metric_values is None:
                    continue
                row_data = {col: base_row[col] for col in ordered_cols}
                for col, (group, param) in metric_map.items():
                    row_data[col] = metric_values.get(group, {}).get(param, pd.NA)

            row_data["Well"] = f"A{well_idx}"
            row_data["Replicate"] = bio_rep
            row_data["BiologicalReplicate"] = bio_rep

            technical = str(base_row.get("TechnicalReplicate", "")).strip()
            row_data["TechnicalReplicate"] = technical if technical else "A"

            output_rows.append(row_data)
            well_idx += 1
            replicated_counts[bio_rep] += 1

    output_datos = pd.DataFrame(output_rows, columns=ordered_cols)
    for int_col in ("Orden", "Replicate", "BiologicalReplicate"):
        output_datos[int_col] = pd.to_numeric(output_datos[int_col], errors="coerce").astype(
            "Int64"
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    backup_path = backup_if_exists(output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        output_datos.to_excel(writer, sheet_name="Datos", index=False)
        base_plot.to_excel(writer, sheet_name="PlotSettings", index=False)

    print(f"Generated file: {output_path}")
    print(f"Total rows in Datos: {len(output_datos)}")
    print(
        "Rows by BiologicalReplicate: "
        + ", ".join(f"{k}={v}" for k, v in replicated_counts.items())
    )

    for bio_rep in (2, 3):
        missing = sorted(base_keys - set(source_by_bio[bio_rep].keys()))
        print(
            f"Missing keys in replicate {bio_rep} source ({len(missing)}): "
            f"{rep2_path if bio_rep == 2 else rep3_path}"
        )
        if missing:
            print(f"  Example missing keys: {missing[:5]}")

    if backup_path is not None:
        print(f"Backup created: {backup_path}")


def main() -> None:
    args = parse_args()
    base_path = Path(args.base).expanduser().resolve()
    rep2_path = Path(args.rep2).expanduser().resolve()
    rep3_path = Path(args.rep3).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    ensure_exists(base_path, "Base platemap")
    ensure_exists(rep2_path, "Replicate 2 workbook")
    ensure_exists(rep3_path, "Replicate 3 workbook")

    build_output(
        base_path=base_path,
        rep2_path=rep2_path,
        rep3_path=rep3_path,
        output_path=output_path,
    )


if __name__ == "__main__":
    main()
